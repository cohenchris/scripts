#!/bin/python3

from __future__ import division             # allows division to be float by default
import csv
import openpyxl
import os
from plexapi.myplex import MyPlexAccount
from plexapi.server import PlexServer
from pprint import pprint
import time
from dotenv import load_dotenv


start_time = time.time()


load_dotenv()
baseurl = os.getenv("PLEX_URL")
token = os.getenv("token")

account = MyPlexAccount(token)
server = PlexServer(baseurl, token)

greenFill = openpyxl.styles.PatternFill("solid", fgColor="CCFFCC")
redFill = openpyxl.styles.PatternFill("solid", fgColor="FFCCCC")

artists = server.library.section("Music").searchArtists()

#   {
#       { "artistName": {
#                           { "albumName": True },
#                           { "albumName": False },
#                           ...
#                       }
#       },
#       { "artistName": {
#                           { "albumName": True },
#                           { "albumName": False },
#                           ...
#                       }
#       },
#       ...
#   }
artist_ratings = {}

# For every artist in the library, check if all of their tracks on each album have been rated
for artist in artists:
    # Add entry for artist in the dictionary
    artist_ratings[artist.title] = {}

# Look through each album from the current artist
    for album in artist.albums():
        # Start album rating as True
        artist_ratings[artist.title][album.title] = True

        # If, for any track on the album, there is no rating, the album has not been completed
        for tracks in album:
            for track in tracks:
                if track.userRating is None:
                    artist_ratings[artist.title][album.title] = False

        # Append each album analysis to the album subarray
        #artist_rating["albums"].append(album_rating)
    # Append each artist rating analysis to the master array
    #artist_ratings.append(artist_rating)


# HEADERS
headers = ["Artist", "Completeness"]
max_albums = len(artist_ratings[max(artist_ratings, default=1, key=lambda x: len(artist_ratings[x]))])
[headers.append(f"Album {str(i)}") for i in range(1, max_albums + 1)]

artist_rating_sheet = [headers]

total_albums = 0
total_albums_rated = 0

# ROWS
# Albums/ratings for every artist
for artist in artist_ratings:
    artist_rating_row = [artist]
    albums_rated = 0
    total_albums = total_albums + len(artist_ratings[artist])

    for album in artist_ratings[artist]:
        if artist_ratings[artist][album]: # True if fully rated, False otherwise
            albums_rated = albums_rated + 1
            total_albums_rated = total_albums_rated + 1

        artist_rating_row.append(f"{album}")

    try:
        percent_rated = f"{round(albums_rated/len(artist_ratings[artist]) * 100, 2)}% ({albums_rated}/{len(artist_ratings[artist])})"
    except ZeroDivisionError:
        print(artist)
        print(artist_ratings[artist])
        print()
        continue

    artist_rating_row.insert(1, percent_rated)
    artist_rating_sheet.append(artist_rating_row)

# Row that adds up totals
artist_rating_sheet.append(["TOTALS", f"{round(total_albums_rated/total_albums * 100, 2)}% ({total_albums_rated}/{total_albums})"])



#
# WRITE ALL CELLS TO SHEET
#

# Start xlsx sheet
wb = openpyxl.Workbook()
sheet = wb.active

for i, row in enumerate(artist_rating_sheet):
    for j, item in enumerate(row):
        coord = openpyxl.utils.get_column_letter(j+1) + str(i+1)
        sheet[coord] = item

        if i == 0:
            # Skip the headers column
            continue

        if (j+1) == 2:  # Completeness column
            if "100.0%" in item:
                # If cell value contains "100.0%", set background color to #CCFFCC
                sheet[coord].fill = greenFill
                pass
            else:
                # If cell value does NOT contain "100.0%", set background color to #FFCCCC
                sheet[coord].fill = redFill
                pass
        elif (j+1) > 2:   # Albums for each artist
            artist = row[0]
            album = row[j]
            if artist_ratings[artist][album]: # True means fully rated
                # If album is fully rated, set background color to #CCFFCC
                sheet[coord].fill = greenFill
                pass
            else:
                # If album is NOT fully rated, set background color to #FFCCCC
                sheet[coord].fill = redFill
                pass

wb.save(filename="musicRatingProgress.xlsx")

#
# SAVE TO NEXTCLOUD (etc/musicRatingProgress.ods)
#

os.system("rclone sync musicRatingProgress.xlsx nextcloud:etc/ --progress")

os.system("rm musicRatingProgress.*")

print("--- %s seconds ---" % (time.time() - start_time))
